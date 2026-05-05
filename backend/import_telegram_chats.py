#!/usr/bin/env python3
"""Build telegram_chats.json from a parent report XLSX file."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_OUTPUT = Path(__file__).with_name("telegram_chats.json")
DEFAULT_CHUNK_SIZE = 7000


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_chat_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = normalize_text(value)
    if text.endswith(".0") and re.fullmatch(r"-?\d+\.0", text):
        text = text[:-2]
    if not re.fullmatch(r"-?\d+", text):
        return ""
    return text


def header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {normalize_text(value).casefold(): index for index, value in enumerate(row) if normalize_text(value)}


def find_column(headers: dict[str, int], *names: str) -> int:
    for name in names:
        key = normalize_text(name).casefold()
        if key in headers:
            return headers[key]
    raise SystemExit(f"Не найдена колонка: {' / '.join(names)}")


def build_chat_config(path: Path, sheet_name: str) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"В файле нет листа {sheet_name!r}. Доступные листы: {', '.join(workbook.sheetnames)}")

    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = header_map(next(rows))
    except StopIteration:
        raise SystemExit("Лист пустой")

    student_col = find_column(headers, "Ученик")
    parent_col = find_column(headers, "Родитель")
    chat_col = find_column(headers, "Чат ID", "chat_id", "chat id")

    students: dict[str, dict[str, set[str]]] = defaultdict(lambda: {"parents": set(), "chat_ids": set()})
    skipped_without_chat: set[str] = set()

    for row in rows:
        student = normalize_text(row[student_col] if student_col < len(row) else "")
        parent = normalize_text(row[parent_col] if parent_col < len(row) else "")
        chat_id = normalize_chat_id(row[chat_col] if chat_col < len(row) else "")

        if not student:
            continue
        if parent:
            students[student]["parents"].add(parent)
        if chat_id:
            students[student]["chat_ids"].add(chat_id)
        else:
            skipped_without_chat.add(student)

    config_students = []
    for student, data in sorted(students.items(), key=lambda item: item[0].casefold()):
        chat_ids = sorted(data["chat_ids"])
        if not chat_ids:
            continue
        config_students.append(
            {
                "name": student,
                "parents": sorted(data["parents"]),
                "chatIds": chat_ids,
                "enabled": True,
            }
        )

    return {
        "students": config_students,
        "meta": {
            "sourceFile": str(path),
            "sourceSheet": sheet_name,
            "studentsTotal": len(students),
            "studentsWithChat": len(config_students),
            "studentsWithoutChat": len(students) - len(config_students),
            "rowsWithoutChat": len(skipped_without_chat),
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert parent chat IDs XLSX to telegram_chats.json")
    parser.add_argument("xlsx", type=Path, help="Path to XLSX file")
    parser.add_argument("--sheet", default="Все", help="Sheet name, default: Все")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help=f"Output JSON, default: {DEFAULT_OUTPUT}")
    parser.add_argument("--print-env", action="store_true", help="Print TELEGRAM_CHATS_JSON_PART_N variables for Railway")
    parser.add_argument("--chunk-size", type=int, default=DEFAULT_CHUNK_SIZE, help=f"Chunk size for --print-env, default: {DEFAULT_CHUNK_SIZE}")
    args = parser.parse_args()

    config = build_chat_config(args.xlsx, args.sheet)
    args.output.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    meta = config["meta"]
    print(f"saved: {args.output}")
    print(f"students total: {meta['studentsTotal']}")
    print(f"students with chat: {meta['studentsWithChat']}")
    print(f"students without chat: {meta['studentsWithoutChat']}")

    if args.print_env:
        compact = json.dumps({"students": config["students"]}, ensure_ascii=False, separators=(",", ":"))
        print()
        print(f"# Add these Railway Variables, do not commit them to GitHub. JSON length: {len(compact)}")
        for index, start in enumerate(range(0, len(compact), args.chunk_size), start=1):
            chunk = compact[start:start + args.chunk_size]
            print(f"TELEGRAM_CHATS_JSON_PART_{index}={chunk}")


if __name__ == "__main__":
    main()
