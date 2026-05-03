#!/usr/bin/env python3
"""Small Soholms proxy/parser backend for the marathon rating site.

The server keeps the Soholms Authorization token on the backend side, downloads
attendance XLSX exports, parses them, and returns JSON rows compatible with the
current frontend rating calculations.
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse
from xml.sax.saxutils import escape as xml_escape

import requests
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Flowable
from reportlab.platypus import Image as ReportImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def load_env_file(path: str) -> None:
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as file:
        for line in file:
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            key, value = stripped.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


load_env_file(os.path.join(os.path.dirname(__file__), ".env"))

API_BASE = os.getenv("SOHOLMS_API_BASE", "https://api.soholms.com").rstrip("/")
DEFAULT_CACHE_SECONDS = int(os.getenv("SOHOLMS_CACHE_SECONDS", "900"))
DEFAULT_CONCURRENCY = int(os.getenv("SOHOLMS_CONCURRENCY", "4"))
MAX_GROUPS_PER_REQUEST = int(os.getenv("SOHOLMS_MAX_GROUPS", "80"))
DEADLINE_SHIFT_DAYS = int(os.getenv("SOHOLMS_DEADLINE_SHIFT_DAYS", "1"))
DEFAULT_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "groups.config.json")
BACKEND_ADMIN_KEY = os.getenv("BACKEND_ADMIN_KEY", "").strip()
DEFAULT_TELEGRAM_CHAT_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "telegram_chats.json")
TELEGRAM_API_BASE = os.getenv("TELEGRAM_API_BASE", "https://api.telegram.org").rstrip("/")

GROUP_TREE_PATH = "/api/v1/learning_group/get_tree"
ATTENDANCE_PATH = "/master/api/learning/attendance-sheet/excel/data"

SUBJECT_LABELS = {
    "математика": "Математика",
    "русский язык": "Русский язык",
    "физика": "Физика",
    "информатика": "Информатика",
    "обществознание": "Обществознание",
    "история": "История",
    "без предмета": "Без предмета",
}

SUBJECT_ALIASES = (
    ("математика", ("матем", "мат11", "м2", "м1")),
    ("русский язык", ("рус", "ря", "русский")),
    ("физика", ("физ", "физика")),
    ("информатика", ("инф", "информ")),
    ("обществознание", ("общ", "обществ")),
    ("история", ("ист", "история")),
)

MONTHS_RU = {
    1: "янв",
    2: "фев",
    3: "мар",
    4: "апр",
    5: "мая",
    6: "июн",
    7: "июл",
    8: "авг",
    9: "сен",
    10: "окт",
    11: "ноя",
    12: "дек",
}

_CACHE: dict[str, tuple[float, Any]] = {}
_PDF_FONT_NAMES: tuple[str, str] | None = None


class BackendError(Exception):
    def __init__(self, message: str, status: int = 500):
        super().__init__(message)
        self.status = status


class PdfIcon(Flowable):
    def __init__(self, kind: str, size: float = 9 * mm):
        super().__init__()
        self.kind = kind
        self.width = size
        self.height = size

    def draw(self) -> None:
        canvas = self.canv
        canvas.saveState()
        canvas.setFillColor(colors.HexColor("#4562f0"))
        canvas.roundRect(0, 0, self.width, self.height, 2.2, fill=1, stroke=0)
        canvas.setStrokeColor(colors.white)
        canvas.setFillColor(colors.white)
        canvas.setLineWidth(1.8)

        if self.kind == "check":
            canvas.line(self.width * 0.27, self.height * 0.52, self.width * 0.43, self.height * 0.35)
            canvas.line(self.width * 0.43, self.height * 0.35, self.width * 0.73, self.height * 0.68)
        else:
            bar_width = self.width * 0.12
            gap = self.width * 0.09
            start = self.width * 0.28
            base = self.height * 0.24
            heights = (self.height * 0.36, self.height * 0.52, self.height * 0.25)
            for index, height in enumerate(heights):
                x = start + index * (bar_width + gap)
                canvas.rect(x, base, bar_width, height, fill=1, stroke=0)

        canvas.restoreState()


@dataclass(frozen=True)
class GroupInfo:
    id: int
    name: str
    subject: str
    teacher: str
    parent_group_ids: tuple[int, ...]
    student_count: int


def clean_authorization_header(value: str) -> str:
    token = value.strip()
    token = re.sub(r"^\s*Authorization:\s*", "", token, flags=re.IGNORECASE).strip()
    token = token.strip("\"'")
    while token.endswith("\\"):
        token = token[:-1].strip()
        token = token.strip("\"'")
    token = token.strip("\"'")
    return token


def get_authorization_header(kind: str = "api") -> str:
    specific_name = "SOHOLMS_EXCEL_TOKEN" if kind == "excel" else "SOHOLMS_API_TOKEN"
    token = clean_authorization_header(os.getenv(specific_name, "")) or clean_authorization_header(os.getenv("SOHOLMS_TOKEN", ""))
    if not token:
        raise BackendError(
            f"{specific_name} or SOHOLMS_TOKEN is not configured",
            HTTPStatus.INTERNAL_SERVER_ERROR,
        )
    return token


def soholms_headers(accept: str = "application/json", kind: str = "api") -> dict[str, str]:
    return {
        "accept": accept,
        "authorization": get_authorization_header(kind),
        "origin": "https://master.soholms.com",
        "referer": "https://master.soholms.com/",
        "user-agent": "marathon-rating-backend/1.0",
    }


def cached(key: str, ttl_seconds: int, loader):
    now = time.time()
    item = _CACHE.get(key)
    if item and now - item[0] < ttl_seconds:
        return item[1]
    try:
        value = loader()
        _CACHE[key] = (now, value)
        return value
    except Exception:
        if item:
            return item[1]
        raise


def clear_cache() -> None:
    _CACHE.clear()


def token_fingerprint(name: str) -> dict[str, Any]:
    value = os.getenv(name, "")
    stripped = clean_authorization_header(value)
    return {
        "configured": bool(stripped),
        "length": len(stripped),
        "startsWithBearer": stripped.lower().startswith("bearer "),
        "preview": f"{stripped[:10]}...{stripped[-6:]}" if len(stripped) > 20 else "",
    }


def request_json(method: str, path: str, **kwargs) -> Any:
    url = f"{API_BASE}{path}"
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            response = requests.request(method, url, timeout=45, **kwargs)
            if response.status_code >= 400:
                raise BackendError(
                    f"Soholms API error {response.status_code}: {response.text[:500]}",
                    response.status_code,
                )
            return response.json()
        except BackendError:
            raise
        except requests.RequestException as error:
            last_error = error
            if attempt < 2:
                time.sleep(1.5 * (attempt + 1))
    raise BackendError(f"Soholms API request failed: {last_error}", HTTPStatus.BAD_GATEWAY)


def fetch_group_tree() -> list[dict[str, Any]]:
    def load():
        data = request_json(
            "POST",
            GROUP_TREE_PATH,
            headers={**soholms_headers(), "content-type": "application/json"},
            data="",
        )
        groups = data.get("learningGroups")
        if not isinstance(groups, list):
            raise BackendError("Unexpected get_tree response: missing learningGroups")
        return groups

    return cached("group_tree", DEFAULT_CACHE_SECONDS, load)


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_group_name(value: Any) -> str:
    return normalize_text(value).casefold()


def compact_group_name(value: Any) -> str:
    return re.sub(r"[^0-9a-zа-яё]+", "", normalize_group_name(value))


def load_group_config() -> dict[str, Any]:
    path = os.getenv("SOHOLMS_GROUP_CONFIG", DEFAULT_CONFIG_PATH)
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as file:
        return json.load(file)


def telegram_chat_ids_from_item(item: dict[str, Any]) -> list[str]:
    raw_chat_ids = item.get("chatIds")
    if raw_chat_ids is None:
        raw_chat_ids = item.get("chat_ids")
    if raw_chat_ids is None:
        raw_chat_ids = [item.get("chatId") or item.get("chat_id")]
    if not isinstance(raw_chat_ids, list):
        raw_chat_ids = [raw_chat_ids]

    chat_ids: list[str] = []
    seen: set[str] = set()
    for raw_chat_id in raw_chat_ids:
        chat_id = normalize_text(raw_chat_id)
        if not chat_id or chat_id in seen:
            continue
        seen.add(chat_id)
        chat_ids.append(chat_id)
    return chat_ids


def load_telegram_chats() -> dict[str, list[str]]:
    path = os.getenv("TELEGRAM_CHAT_CONFIG", DEFAULT_TELEGRAM_CHAT_CONFIG_PATH)
    raw_json = os.getenv("TELEGRAM_CHATS_JSON", "").strip()
    if raw_json:
        raw = json.loads(raw_json)
    elif os.path.exists(path):
        with open(path, "r", encoding="utf-8") as file:
            raw = json.load(file)
    else:
        return {}

    if isinstance(raw, dict) and isinstance(raw.get("students"), list):
        items = raw["students"]
    elif isinstance(raw, list):
        items = raw
    elif isinstance(raw, dict):
        items = [{"name": name, "chatId": chat_id} for name, chat_id in raw.items()]
    else:
        raise BackendError("Unexpected TELEGRAM_CHAT_CONFIG format", HTTPStatus.INTERNAL_SERVER_ERROR)

    chats: dict[str, list[str]] = {}
    for item in items:
        if not isinstance(item, dict) or item.get("enabled") is False:
            continue
        name = normalize_text(item.get("name"))
        chat_ids = telegram_chat_ids_from_item(item)
        if name and chat_ids:
            current = chats.setdefault(normalize_group_name(name), [])
            for chat_id in chat_ids:
                if chat_id not in current:
                    current.append(chat_id)
    return chats


def children_by_parent(groups: list[dict[str, Any]]) -> dict[int, list[int]]:
    children: dict[int, list[int]] = defaultdict(list)
    for group in groups:
        group_id = group.get("id")
        if not isinstance(group_id, int):
            continue
        for parent_id in group.get("parentGroupIds") or []:
            if isinstance(parent_id, int):
                children[parent_id].append(group_id)
    return children


def descendant_ids(root_ids: set[int], groups: list[dict[str, Any]]) -> set[int]:
    children = children_by_parent(groups)
    result: set[int] = set()
    stack = list(root_ids)
    while stack:
        group_id = stack.pop()
        for child_id in children.get(group_id, []):
            if child_id in result:
                continue
            result.add(child_id)
            stack.append(child_id)
    return result


def similar_group_names(name: str, groups: list[dict[str, Any]], limit: int = 6) -> list[str]:
    target = compact_group_name(name)
    if not target:
        return []

    scored: list[tuple[float, str]] = []
    seen: set[str] = set()
    for group in groups:
        group_name = normalize_text(group.get("name"))
        if not group_name or group_name in seen:
            continue
        seen.add(group_name)
        compact = compact_group_name(group_name)
        score = SequenceMatcher(None, target, compact).ratio()
        if target in compact or compact in target:
            score += 0.35
        scored.append((score, group_name))

    return [name for score, name in sorted(scored, reverse=True)[:limit] if score >= 0.45]


def resolve_config_group_ids(groups: list[dict[str, Any]], config: dict[str, Any]) -> tuple[set[int], list[str], dict[str, list[str]]]:
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    existing_ids = {group.get("id") for group in groups if isinstance(group.get("id"), int)}
    for group in groups:
        by_name[normalize_group_name(group.get("name"))].append(group)

    selected: set[int] = set()
    missing: list[str] = []
    candidates: dict[str, list[str]] = {}

    for raw_name in config.get("groupNames") or []:
        matches = by_name.get(normalize_group_name(raw_name), [])
        if not matches:
            missing_name = str(raw_name)
            missing.append(missing_name)
            candidates[missing_name] = similar_group_names(missing_name, groups)
            continue
        selected.update(group["id"] for group in matches if isinstance(group.get("id"), int))

    descendant_roots: set[int] = set()
    for raw_name in config.get("includeDescendantsOfNames") or []:
        matches = by_name.get(normalize_group_name(raw_name), [])
        if not matches:
            missing_name = str(raw_name)
            missing.append(missing_name)
            candidates[missing_name] = similar_group_names(missing_name, groups)
            continue
        descendant_roots.update(group["id"] for group in matches if isinstance(group.get("id"), int))
    selected.update(descendant_ids(descendant_roots, groups))

    for group_id in config.get("groupIds") or []:
        try:
            numeric_id = int(group_id)
        except (TypeError, ValueError):
            missing_name = str(group_id)
            missing.append(missing_name)
            candidates[missing_name] = []
            continue
        if numeric_id not in existing_ids:
            missing_name = f"groupId:{numeric_id}"
            missing.append(missing_name)
            candidates[missing_name] = []
            continue
        selected.add(numeric_id)

    return selected, missing, candidates


def teacher_name(teacher: dict[str, Any]) -> str:
    parts = [
        normalize_text(teacher.get("lastName")),
        normalize_text(teacher.get("firstName")),
        normalize_text(teacher.get("middleName")),
    ]
    return " ".join([part for part in parts if part])


def group_teacher_label(group: dict[str, Any]) -> str:
    names: list[str] = []
    seen: set[str] = set()
    for teacher in group.get("teachers") or []:
        if teacher.get("isDisabled"):
            continue
        name = teacher_name(teacher)
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    return ", ".join(names) or "Без преподавателя"


def infer_subject(value: str, parent_subject: str = "") -> str:
    text = f"{value} {parent_subject}".lower()
    for subject, aliases in SUBJECT_ALIASES:
        if any(alias in text for alias in aliases):
            return subject
    return "без предмета"


def infer_level(value: str) -> str:
    text = value.upper()
    if "ОГЭ" in text:
        return "ОГЭ"
    if "ЕГЭ" in text:
        return "ЕГЭ"
    if re.search(r"(^|[^0-9])9([^0-9]|$)", text):
        return "ОГЭ"
    if re.search(r"(^|[^0-9])11([^0-9]|$)", text):
        return "ЕГЭ"
    return ""


def discipline_matches_group(discipline: Any, group: GroupInfo) -> bool:
    text = normalize_text(discipline)
    lowered = text.casefold()
    if not text:
        return True
    if "основн" in lowered:
        return False

    discipline_subject = infer_subject(text)
    if group.subject != "без предмета" and discipline_subject != "без предмета" and discipline_subject != group.subject:
        return False

    expected_level = infer_level(group.name)
    discipline_level = infer_level(text)
    if expected_level and discipline_level and discipline_level != expected_level:
        return False

    return True


def build_group_index(groups: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    result = {}
    for group in groups:
        group_id = group.get("id")
        if isinstance(group_id, int):
            result[group_id] = group
    return result


def parent_subject_name(group: dict[str, Any], by_id: dict[int, dict[str, Any]]) -> str:
    for parent_id in group.get("parentGroupIds") or []:
        parent = by_id.get(parent_id)
        if parent:
            subject = infer_subject(parent.get("name", ""))
            if subject != "без предмета":
                return subject
    return ""


def selected_groups(
    group_ids: set[int] | None = None,
    subjects: set[str] | None = None,
    include_virtual: bool = False,
    origins: set[str] | None = None,
) -> list[GroupInfo]:
    groups = fetch_group_tree()
    by_id = build_group_index(groups)
    result: list[GroupInfo] = []

    for group in groups:
        group_id = group.get("id")
        if not isinstance(group_id, int):
            continue
        if group_ids is not None and group_id not in group_ids:
            continue
        if not include_virtual and group.get("purpose") != "PhysicalLearning":
            continue
        if origins and normalize_text(group.get("origin")) not in origins:
            continue
        if not group.get("studentIds"):
            continue

        teacher = group_teacher_label(group)
        if teacher == "Без преподавателя" and group_ids is None:
            continue

        parent_subject = parent_subject_name(group, by_id)
        subject = infer_subject(group.get("name", ""), parent_subject)
        if subjects and subject not in subjects:
            continue

        result.append(
            GroupInfo(
                id=group_id,
                name=normalize_text(group.get("name")),
                subject=subject,
                teacher=teacher,
                parent_group_ids=tuple(group.get("parentGroupIds") or []),
                student_count=len(group.get("studentIds") or []),
            )
        )

    return sorted(result, key=lambda item: (item.subject, item.name, item.id))


def searchable_group_row(group: dict[str, Any], by_id: dict[int, dict[str, Any]]) -> dict[str, Any]:
    parent_subject = parent_subject_name(group, by_id)
    return {
        "id": group.get("id"),
        "name": normalize_text(group.get("name")),
        "subject": infer_subject(group.get("name", ""), parent_subject),
        "teacher": group_teacher_label(group),
        "purpose": group.get("purpose"),
        "origin": group.get("origin"),
        "parent_group_ids": group.get("parentGroupIds") or [],
        "student_count": len(group.get("studentIds") or []),
    }


def search_groups(search: str) -> list[dict[str, Any]]:
    groups = fetch_group_tree()
    by_id = build_group_index(groups)
    query = normalize_group_name(search)
    compact_query = compact_group_name(search)
    result: list[dict[str, Any]] = []

    for group in groups:
        name = group.get("name")
        if not normalize_text(name):
            continue
        normalized_name = normalize_group_name(name)
        compact_name = compact_group_name(name)
        if query in normalized_name or compact_query in compact_name:
            result.append(searchable_group_row(group, by_id))

    return sorted(result, key=lambda item: (item["subject"], item["name"], item["id"] or 0))


def fetch_attendance_xlsx(group_id: int, period_from: str, period_to: str) -> bytes:
    params = {
        "learningGroupId": group_id,
        "periodFrom": period_from,
        "periodTo": period_to,
        "isInteractiveLessons": "true",
        "isAcademicLessons": "true",
        "isStudentsOutOfPeriod": "false",
        "withHomeworks": "true",
        "academicDisciplineIds": "",
        "masterClientIds": "",
    }
    url = f"{API_BASE}{ATTENDANCE_PATH}?{urlencode(params)}"
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            response = requests.get(
                url,
                headers=soholms_headers(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
                    kind="excel",
                ),
                timeout=90,
            )
            if response.status_code >= 400:
                raise BackendError(
                    f"Soholms XLSX error {response.status_code} for group {group_id}: {response.text[:500]}",
                    response.status_code,
                )
            return response.content
        except BackendError:
            raise
        except requests.RequestException as error:
            last_error = error
            if attempt < 2:
                time.sleep(1.5 * (attempt + 1))
    raise BackendError(f"Soholms XLSX request failed for group {group_id}: {last_error}", HTTPStatus.BAD_GATEWAY)


def score_value(*values: Any) -> float | None:
    for value in values:
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str) and value.strip():
            match = re.match(r"^\s*(\d+(?:[,.]\d+)?)\s*/\s*(\d+(?:[,.]\d+)?)\s*$", value)
            if match:
                got = float(match.group(1).replace(",", "."))
                total = float(match.group(2).replace(",", "."))
                return 0.0 if total == 0 else got / total * 100
            try:
                return float(value.replace(",", "."))
            except ValueError:
                continue
    return None


def iso_date(value: Any, shift_days: int = 0) -> str:
    if isinstance(value, datetime):
        return (value.date() + timedelta(days=shift_days)).isoformat()
    if isinstance(value, date):
        return (value + timedelta(days=shift_days)).isoformat()
    return normalize_text(value)


def date_label(value: Any, shift_days: int = 0) -> str:
    if isinstance(value, datetime):
        d = value.date() + timedelta(days=shift_days)
    elif isinstance(value, date):
        d = value + timedelta(days=shift_days)
    else:
        return normalize_text(value)
    return f"{d.day:02d}.{MONTHS_RU.get(d.month, str(d.month))}"


def parse_lesson_number(value: Any) -> int:
    match = re.match(r"^\s*(\d+)\.", normalize_text(value))
    return int(match.group(1)) if match else 0


def is_day_lesson(value: Any) -> bool:
    return "день" in normalize_text(value).lower()


def lesson_day_key(lesson: Any, lesson_date: Any) -> str:
    day_order = parse_lesson_number(lesson)
    if day_order:
        return f"day:{day_order}"
    date_key = iso_date(lesson_date)
    if date_key:
        return f"date:{date_key}"
    return f"lesson:{normalize_text(lesson)}"


def header_index(headers: list[Any], *names: str, default: int | None = None) -> int | None:
    normalized_headers = [normalize_text(header).casefold() for header in headers]
    for name in names:
        normalized_name = normalize_text(name).casefold()
        for index, header in enumerate(normalized_headers):
            if header == normalized_name:
                return index
    return default


def row_value(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def late_days(lesson_date: Any, submitted_at: Any) -> int:
    if not isinstance(lesson_date, (date, datetime)) or not isinstance(submitted_at, (date, datetime)):
        return 0

    lesson_day = lesson_date.date() if isinstance(lesson_date, datetime) else lesson_date
    submitted_day = submitted_at.date() if isinstance(submitted_at, datetime) else submitted_at
    due_day = lesson_day + timedelta(days=DEADLINE_SHIFT_DAYS)
    return max(0, (submitted_day - due_day).days)


def late_penalty(lesson_date: Any, submitted_at: Any) -> int:
    return 1 if late_days(lesson_date, submitted_at) > 0 else 0


def parse_attendance_xlsx(content: bytes, group: GroupInfo) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=3, max_row=3))]
    columns = {
        "student_id": header_index(headers, "ID ученика", default=0),
        "name": header_index(headers, "Ученик", default=1),
        "group": header_index(headers, "Учебная группа", default=2),
        "discipline": header_index(headers, "Дисциплина", default=3),
        "lesson": header_index(headers, "Урок", default=4),
        "lesson_date": header_index(headers, "Дата урока", default=5),
        "lesson_score": header_index(headers, "Оценка за урок", default=8),
        "homework_score": header_index(headers, "Оценка за ДЗ", default=9),
        "checkpoint_score": header_index(headers, "Оценка за СР", default=10),
        "control_score": header_index(headers, "Оценка за КР", default=11),
        "assignment_status": header_index(headers, "Статус сдачи", default=17),
        "assignment_score": header_index(headers, "Оценка", default=18),
        "submitted_at": header_index(headers, "Дата сдачи", default=19),
    }
    students: dict[str, dict[str, Any]] = {}
    group_day_keys: list[str] = []
    group_day_key_set: set[str] = set()
    day_key_orders: dict[str, int] = {}
    current_day: dict[str, Any] | None = None

    def register_group_day(lesson: Any, lesson_date: Any) -> tuple[str, int]:
        raw_day_order = parse_lesson_number(lesson)
        day_key = lesson_day_key(lesson, lesson_date)
        if day_key not in group_day_key_set:
            group_day_key_set.add(day_key)
            group_day_keys.append(day_key)
            day_key_orders[day_key] = raw_day_order or len(group_day_keys)
        return day_key, day_key_orders[day_key]

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        student_id = row_value(row, columns["student_id"])
        name = row_value(row, columns["name"])
        xlsx_group = row_value(row, columns["group"])
        discipline = row_value(row, columns["discipline"])
        lesson = row_value(row, columns["lesson"])
        lesson_date = row_value(row, columns["lesson_date"])
        lesson_score = row_value(row, columns["lesson_score"])
        homework_score = row_value(row, columns["homework_score"])
        checkpoint_score = row_value(row, columns["checkpoint_score"])
        control_score = row_value(row, columns["control_score"])
        assignment_status = row_value(row, columns["assignment_status"])
        assignment_score = row_value(row, columns["assignment_score"])
        submitted_at = row_value(row, columns["submitted_at"])
        score = score_value(checkpoint_score, control_score, homework_score, lesson_score)
        is_potential_day = is_day_lesson(lesson) or isinstance(lesson_date, (date, datetime))
        is_scored_day = score is not None and is_potential_day

        if not name:
            if current_day and (normalize_text(assignment_status) or assignment_score not in (None, "")):
                lesson_late_days = late_penalty(current_day.get("lessonDate"), submitted_at)
                late_by_lesson = current_day["item"].setdefault("_lateDaysByLesson", {})
                lesson_key = current_day["dayOrder"]
                previous = late_by_lesson.get(lesson_key)
                if previous is None or lesson_late_days < previous:
                    late_by_lesson[lesson_key] = lesson_late_days
                current_day["dailyScore"]["lateDays"] = lesson_late_days
            continue

        if current_day and not is_day_lesson(lesson) and (normalize_text(assignment_status) or assignment_score not in (None, "")):
            lesson_late_days = late_penalty(current_day.get("lessonDate"), submitted_at)
            late_by_lesson = current_day["item"].setdefault("_lateDaysByLesson", {})
            lesson_key = current_day["dayOrder"]
            previous = late_by_lesson.get(lesson_key)
            if previous is None or lesson_late_days < previous:
                late_by_lesson[lesson_key] = lesson_late_days
                current_day["dailyScore"]["lateDays"] = lesson_late_days
            continue

        if not discipline_matches_group(discipline, group):
            current_day = None
            continue

        day_key = ""
        day_order = 0
        if is_potential_day:
            day_key, day_order = register_group_day(lesson, lesson_date)

        if not name or not is_scored_day:
            current_day = None
            continue

        student_key = str(int(student_id)) if isinstance(student_id, (int, float)) else normalize_text(name)
        item = students.setdefault(
            student_key,
            {
                "subject": group.subject if group.subject != "без предмета" else infer_subject(discipline or xlsx_group or group.name),
                "level": infer_level(discipline) or infer_level(group.name),
                "group": normalize_text(xlsx_group) or group.name,
                "name": normalize_text(name),
                "teacher": group.teacher,
                "dailyScores": [],
                "_lateDaysByLesson": {},
            },
        )
        daily_score = {
            "dateKey": iso_date(lesson_date, shift_days=DEADLINE_SHIFT_DAYS),
            "dateLabel": date_label(lesson_date, shift_days=DEADLINE_SHIFT_DAYS),
            "dateOrder": day_order,
            "dayKey": day_key,
            "score": score,
            "lateDays": 0,
        }
        item["dailyScores"].append(daily_score)
        current_day = {
            "item": item,
            "dayOrder": day_order,
            "lessonDate": lesson_date,
            "dailyScore": daily_score,
        }

    rows: list[dict[str, Any]] = []
    group_days_total = len(group_day_keys)
    for item in students.values():
        daily_scores = item["dailyScores"]
        scores_by_day: dict[str, float] = {}
        for day in daily_scores:
            day_key = str(day.get("dayKey") or day.get("dateOrder") or "")
            if not day_key:
                continue
            scores_by_day[day_key] = max(scores_by_day.get(day_key, 0.0), float(day["score"]))
        scores = list(scores_by_day.values())
        days_done = len(scores_by_day)
        days_total = group_days_total or days_done
        quality = sum(scores) / days_done if days_done else 0.0
        coefficient = days_done / days_total if days_total else 0.0
        base_score = quality * coefficient
        penalty = float(sum(item.get("_lateDaysByLesson", {}).values()))
        final_score = max(0.0, base_score - penalty)
        item.pop("_lateDaysByLesson", None)
        for day in daily_scores:
            day.pop("dayKey", None)

        item.update(
            {
                "daysDone": days_done,
                "daysTotal": days_total,
                "coefficient": coefficient,
                "quality": quality,
                "baseScore": base_score,
                "penalty": penalty,
                "finalScore": final_score,
                "score": final_score,
            }
        )
        rows.append(item)

    return rows


def inspect_attendance_xlsx(content: bytes, sample_limit: int = 12) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [normalize_text(cell.value) for cell in next(worksheet.iter_rows(min_row=3, max_row=3))]
    columns = {
        "name": header_index(headers, "Ученик", default=1),
        "lesson": header_index(headers, "Урок", default=4),
        "lesson_date": header_index(headers, "Дата урока", default=5),
        "assignment_status": header_index(headers, "Статус сдачи", default=17),
        "assignment_score": header_index(headers, "Оценка", default=18),
        "submitted_at": header_index(headers, "Дата сдачи", default=19),
    }
    total_rows = 0
    day_rows = 0
    assignment_rows = 0
    submitted_rows = 0
    named_assignment_rows = 0
    unnamed_assignment_rows = 0
    samples: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        total_rows += 1
        name = row_value(row, columns["name"])
        lesson = row_value(row, columns["lesson"])
        status = row_value(row, columns["assignment_status"])
        assignment_score = row_value(row, columns["assignment_score"])
        submitted_at = row_value(row, columns["submitted_at"])
        has_assignment = bool(normalize_text(status)) or assignment_score not in (None, "")

        if is_day_lesson(lesson):
            day_rows += 1
        if has_assignment:
            assignment_rows += 1
            if name:
                named_assignment_rows += 1
            else:
                unnamed_assignment_rows += 1
        if submitted_at:
            submitted_rows += 1
        if has_assignment and len(samples) < sample_limit:
            samples.append({
                "name": normalize_text(name),
                "lesson": normalize_text(lesson),
                "lessonDate": iso_date(row_value(row, columns["lesson_date"])),
                "status": normalize_text(status),
                "assignmentScore": normalize_text(assignment_score),
                "submittedAt": iso_date(submitted_at),
            })

    return {
        "headers": headers,
        "columns": columns,
        "stats": {
            "totalRows": total_rows,
            "dayRows": day_rows,
            "assignmentRows": assignment_rows,
            "submittedRows": submitted_rows,
            "namedAssignmentRows": named_assignment_rows,
            "unnamedAssignmentRows": unnamed_assignment_rows,
        },
        "samples": samples,
    }


def add_places(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_group: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    by_school: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)

    for row in rows:
        by_group[(row.get("subject", ""), row.get("level", ""), row.get("group", ""))].append(row)
        by_school[(row.get("subject", ""), row.get("level", ""))].append(row)

    def apply_place(items: list[dict[str, Any]], key: str) -> None:
        sorted_items = sorted(items, key=lambda row: float(row.get("finalScore") or 0), reverse=True)
        previous_score = None
        place = 0
        for item in sorted_items:
            score = float(item.get("finalScore") or 0)
            if score != previous_score:
                place += 1
                previous_score = score
            item[key] = place

    for items in by_group.values():
        apply_place(items, "groupPlace")
    for items in by_school.values():
        apply_place(items, "schoolPlace")

    return rows


def public_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "subject": row.get("subject", ""),
        "level": row.get("level", ""),
        "name": row.get("name", ""),
        "teacher": row.get("teacher", "Без преподавателя"),
        "score": row.get("score", 0),
        "finalScore": row.get("finalScore", row.get("score", 0)),
        "groupPlace": row.get("groupPlace", 0),
        "schoolPlace": row.get("schoolPlace", 0),
    }


def strip_for_public(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        **payload,
        "rows": [public_row(row) for row in payload.get("rows", [])],
    }


def format_report_number(value: Any, digits: int = 2) -> str:
    return f"{float(value or 0):.{digits}f}".replace(".", ",")


def average_values(values: list[Any]) -> float:
    numbers = [float(value or 0) for value in values]
    return sum(numbers) / len(numbers) if numbers else 0.0


def register_pdf_fonts() -> tuple[str, str]:
    global _PDF_FONT_NAMES
    if _PDF_FONT_NAMES:
        return _PDF_FONT_NAMES

    regular_candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/Library/Fonts/DejaVuSans.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/Library/Fonts/Arial.ttf",
    )
    bold_candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/Library/Fonts/DejaVuSans-Bold.ttf",
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        "/Library/Fonts/Arial Bold.ttf",
    )
    regular_path = next((path for path in regular_candidates if os.path.exists(path)), "")
    bold_path = next((path for path in bold_candidates if os.path.exists(path)), "")

    if regular_path:
        pdfmetrics.registerFont(TTFont("MarathonSans", regular_path))
        regular_font = "MarathonSans"
    else:
        regular_font = "Helvetica"

    if bold_path:
        pdfmetrics.registerFont(TTFont("MarathonSans-Bold", bold_path))
        bold_font = "MarathonSans-Bold"
    else:
        bold_font = regular_font if regular_path else "Helvetica-Bold"

    _PDF_FONT_NAMES = (regular_font, bold_font)
    return _PDF_FONT_NAMES


def pdf_text(value: Any) -> Paragraph:
    regular_font, _ = register_pdf_fonts()
    style = ParagraphStyle(
        "Cell",
        fontName=regular_font,
        fontSize=7.2,
        leading=8.6,
        textColor=colors.HexColor("#20242a"),
    )
    return Paragraph(xml_escape(normalize_text(value)), style)


def pdf_paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    text = str(value or "").strip()
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    return Paragraph(xml_escape("\n".join(lines)).replace("\n", "<br/>"), style)


def find_logo_path() -> str:
    candidates = (
        os.path.join(os.path.dirname(__file__), "logoplanka-cropped.png"),
        os.path.join(os.path.dirname(__file__), "logoplanka.png"),
        os.path.join(os.path.dirname(__file__), "..", "logoplanka.png"),
        os.path.join(os.getcwd(), "logoplanka.png"),
    )
    return next((path for path in candidates if os.path.exists(path)), "")


def build_student_pdf_report(student_name: str, rows: list[dict[str, Any]], period: dict[str, Any]) -> bytes:
    regular_font, bold_font = register_pdf_fonts()
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("subject") or ""),
            str(row.get("level") or ""),
            str(row.get("group") or ""),
        ),
    )

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    title_white = ParagraphStyle(
        "TitleWhite",
        fontName=bold_font,
        fontSize=25,
        leading=29,
        textColor=colors.white,
    )
    subtitle_white = ParagraphStyle(
        "SubtitleWhite",
        fontName=regular_font,
        fontSize=12.2,
        leading=15,
        textColor=colors.white,
    )
    body_style = ParagraphStyle(
        "ParentBody",
        fontName=regular_font,
        fontSize=12.7,
        leading=17,
        textColor=colors.HexColor("#2d3146"),
    )
    body_bold = ParagraphStyle(
        "ParentBodyBold",
        parent=body_style,
        fontName=bold_font,
        fontSize=13.2,
        leading=17,
    )
    small_label = ParagraphStyle(
        "SmallLabel",
        fontName=regular_font,
        fontSize=8.4,
        leading=11,
        textColor=colors.HexColor("#8b97ab"),
    )
    student_name_style = ParagraphStyle(
        "StudentName",
        fontName=bold_font,
        fontSize=13.2,
        leading=16,
        textColor=colors.HexColor("#2d3146"),
    )
    section_title = ParagraphStyle(
        "SectionTitle",
        fontName=bold_font,
        fontSize=14.3,
        leading=17,
        textColor=colors.HexColor("#2d3146"),
    )
    table_header_style = ParagraphStyle(
        "TableHeader",
        fontName=bold_font,
        fontSize=8.8,
        leading=10.5,
        alignment=1,
        textColor=colors.white,
    )
    table_cell_style = ParagraphStyle(
        "TableCell",
        fontName=regular_font,
        fontSize=8.8,
        leading=10.8,
        alignment=1,
        textColor=colors.HexColor("#2d3146"),
    )
    table_cell_bold = ParagraphStyle(
        "TableCellBold",
        parent=table_cell_style,
        fontName=bold_font,
    )

    logo_path = find_logo_path()
    if logo_path:
        logo_cell: Any = ReportImage(logo_path, width=22 * mm, height=25 * mm)
    else:
        logo_cell = pdf_paragraph("постоянная\nпланка", ParagraphStyle("LogoFallback", fontName=bold_font, fontSize=8, leading=9, alignment=1))
    logo_box = Table([[logo_cell]], colWidths=[26 * mm], rowHeights=[26 * mm])
    logo_box.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#4562f0")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    header = Table(
        [
            [
                logo_box,
                [
                    Paragraph("ОТЧЕТ ПО МАРАФОНАМ", title_white),
                    Paragraph("Ежедневная отработка первой части ЕГЭ/ОГЭ", subtitle_white),
                ],
            ]
        ],
        colWidths=[34 * mm, 135 * mm],
        rowHeights=[32 * mm],
    )
    header.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#4562f0")),
                ("ALIGN", (0, 0), (0, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (0, 0), 4 * mm),
                ("RIGHTPADDING", (0, 0), (0, 0), 2 * mm),
                ("TOPPADDING", (0, 0), (0, 0), 3 * mm),
                ("BOTTOMPADDING", (0, 0), (0, 0), 3 * mm),
                ("LEFTPADDING", (1, 0), (1, 0), 5 * mm),
                ("RIGHTPADDING", (1, 0), (1, 0), 8 * mm),
            ]
        )
    )
    story: list[Any] = [header, Spacer(1, 6 * mm)]

    about = Table(
        [
            [pdf_paragraph("Что такое марафон", body_bold)],
            [
                pdf_paragraph(
                    "Марафон - это ежедневные 30-минутные подборки заданий первой части ЕГЭ/ОГЭ с "
                    "автоматической проверкой. Все задания взяты из банка ФИПИ и помогают системно "
                    "закрыть пробелы и повысить итоговый балл.",
                    body_style,
                )
            ],
        ],
        colWidths=[169 * mm],
    )
    about.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f2f6ff")),
                ("LINEBEFORE", (0, 0), (0, -1), 3, colors.HexColor("#4562f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
                ("TOPPADDING", (0, 1), (-1, 1), 0),
                ("BOTTOMPADDING", (0, 1), (-1, 1), 12),
            ]
        )
    )
    story.extend([about, Spacer(1, 6 * mm)])

    check = Table(
        [[PdfIcon("check"), pdf_paragraph("В формате марафона ученики получают", body_bold)]],
        colWidths=[9 * mm, 153 * mm],
        rowHeights=[9 * mm],
    )
    check.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 8),
            ]
        )
    )
    story.append(check)
    bullet_text = (
        "- ежедневные задания с напоминаниями\n"
        "- еженедельную статистику и рейтинги по группе, предмету и школе\n"
        "- карту ошибок по итогам прохождения\n"
        "- при необходимости разбор сложных заданий (по математике ЕГЭ - с видеоразборами)\n"
        "- возможность задать вопросы преподавателю\n"
        "- систему мотивации с призами за регулярную работу"
    )
    story.extend(
        [
            Spacer(1, 4 * mm),
            pdf_paragraph(bullet_text, body_style),
            Spacer(1, 2.2 * mm),
            pdf_paragraph(
                "Такой формат помогает поддерживать темп подготовки и значительно повышает результат "
                "по первой части экзамена.",
                body_style,
            ),
            Spacer(1, 5 * mm),
            pdf_paragraph("Ученик", small_label),
            pdf_paragraph(student_name, student_name_style),
            Spacer(1, 4 * mm),
        ]
    )

    section = Table(
        [[PdfIcon("chart"), pdf_paragraph("Статистика по марафонам", section_title)]],
        colWidths=[9 * mm, 153 * mm],
        rowHeights=[9 * mm],
    )
    section.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 8),
            ]
        )
    )
    story.extend([section, Spacer(1, 4 * mm)])

    table_data: list[list[Any]] = [
        [
            pdf_paragraph("Предмет", table_header_style),
            pdf_paragraph("Дней сделано", table_header_style),
            pdf_paragraph("Дней всего", table_header_style),
            pdf_paragraph("Качество", table_header_style),
            pdf_paragraph("Качество макс", table_header_style),
            pdf_paragraph("Преподаватель", table_header_style),
        ]
    ]
    for row in sorted_rows:
        subject = SUBJECT_LABELS.get(str(row.get("subject") or ""), row.get("subject") or "")
        table_data.append(
            [
                pdf_paragraph(str(subject).lower(), table_cell_bold),
                pdf_paragraph(int(row.get("daysDone") or 0), table_cell_style),
                pdf_paragraph(int(row.get("daysTotal") or 0), table_cell_style),
                pdf_paragraph(format_report_number(row.get("quality"), 0), table_cell_style),
                pdf_paragraph("100", table_cell_style),
                pdf_paragraph(row.get("teacher") or "Без преподавателя", table_cell_style),
            ]
        )

    col_widths = [22 * mm, 32 * mm, 28 * mm, 27 * mm, 32 * mm, 28 * mm]
    report_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    report_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4562f0")),
                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#d9e5ea")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.extend([report_table, Spacer(1, 8 * mm)])

    parent_note = Table(
        [
            [
                pdf_paragraph(
                    "Со своей стороны мы ежедневно отслеживаем участие ребят, напоминаем о заданиях, "
                    "ведём рейтинги и проводим конкурс с призами для самых активных участников. Но "
                    "практика показывает, что при поддержке родителей результаты марафона становятся "
                    "значительно выше.\n\n"
                    "Будем очень благодарны, если вы сможете уточнять у ребёнка, выполняет ли он "
                    "задания марафона - такая вовлечённость заметно помогает сохранять регулярность и "
                    "повышает эффективность подготовки.",
                    body_style,
                )
            ]
        ],
        colWidths=[169 * mm],
    )
    parent_note.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff7ed")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#ffbd3e")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ]
        )
    )
    story.append(parent_note)
    document.build(story)
    return buffer.getvalue()


def build_student_telegram_report(student_name: str, rows: list[dict[str, Any]], period: dict[str, Any]) -> str:
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("subject") or ""),
            str(row.get("level") or ""),
            str(row.get("group") or ""),
        ),
    )
    period_label = f"{period.get('from', '...')} — {period.get('to', '...')}"
    average_coefficient = average_values([row.get("coefficient") for row in sorted_rows])
    average_quality = average_values([row.get("quality") for row in sorted_rows])
    average_final = average_values([row.get("finalScore") for row in sorted_rows])
    total_penalty = sum(float(row.get("penalty") or 0) for row in sorted_rows)

    lines = [
        f"Отчёт по марафону: {student_name}",
        f"Период: {period_label}",
        "",
        f"Коэффициент: {format_report_number(average_coefficient)}",
        f"Качество: {format_report_number(average_quality)}",
        f"Штраф: {format_report_number(total_penalty, 0)}",
        f"Итоговый балл: {format_report_number(average_final)}",
        "",
    ]

    for row in sorted_rows:
        subject = SUBJECT_LABELS.get(str(row.get("subject") or ""), row.get("subject") or "")
        lines.extend([
            f"{subject} {row.get('level') or ''} · {row.get('group') or ''}",
            f"Дни: {int(row.get('daysDone') or 0)}/{int(row.get('daysTotal') or 0)} · "
            f"Качество: {format_report_number(row.get('quality'))} · "
            f"Балл: {format_report_number(row.get('baseScore'))} · "
            f"Штраф: {format_report_number(row.get('penalty'), 0)} · "
            f"Итог: {format_report_number(row.get('finalScore'))}",
            f"Место в группе: {int(row.get('groupPlace') or 0)} · "
            f"Место в школе: {int(row.get('schoolPlace') or 0)}",
            f"Преподаватель: {row.get('teacher') or 'Без преподавателя'}",
            "",
        ])

    text = "\n".join(lines).strip()
    if len(text) > 3900:
        return text[:3850].rstrip() + "\n\nОтчёт сокращён, полная версия доступна в админке."
    return text


def report_filename(student_name: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-zА-Яа-яЁё_-]+", "_", normalize_text(student_name)).strip("_")
    return f"marathon_report_{cleaned or 'student'}.pdf"


def send_telegram_message(chat_id: str, text: str) -> dict[str, Any]:
    token = clean_authorization_header(os.getenv("TELEGRAM_BOT_TOKEN", ""))
    if not token:
        raise BackendError("TELEGRAM_BOT_TOKEN is not configured", HTTPStatus.INTERNAL_SERVER_ERROR)
    response = requests.post(
        f"{TELEGRAM_API_BASE}/bot{token}/sendMessage",
        json={
            "chat_id": chat_id,
            "text": text,
            "disable_web_page_preview": True,
        },
        timeout=30,
    )
    if response.status_code >= 400:
        raise BackendError(f"Telegram API error {response.status_code}: {response.text[:500]}", response.status_code)
    return response.json()


def send_telegram_document(chat_id: str, pdf_bytes: bytes, filename: str, caption: str) -> dict[str, Any]:
    token = clean_authorization_header(os.getenv("TELEGRAM_BOT_TOKEN", ""))
    if not token:
        raise BackendError("TELEGRAM_BOT_TOKEN is not configured", HTTPStatus.INTERNAL_SERVER_ERROR)
    response = requests.post(
        f"{TELEGRAM_API_BASE}/bot{token}/sendDocument",
        data={
            "chat_id": chat_id,
            "caption": caption[:1024],
        },
        files={
            "document": (filename, pdf_bytes, "application/pdf"),
        },
        timeout=60,
    )
    if response.status_code >= 400:
        raise BackendError(f"Telegram API error {response.status_code}: {response.text[:500]}", response.status_code)
    return response.json()


def send_telegram_reports(
    rows: list[dict[str, Any]],
    period: dict[str, Any],
    student_name: str = "",
    *,
    send_pdf: bool = False,
) -> dict[str, Any]:
    chats = load_telegram_chats()
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        name = normalize_text(row.get("name"))
        if student_name and name != student_name:
            continue
        if name:
            grouped[name].append(row)

    sent: list[dict[str, Any]] = []
    missing: list[str] = []
    errors: list[dict[str, Any]] = []

    for name in sorted(grouped):
        chat_ids = chats.get(normalize_group_name(name)) or []
        if not chat_ids:
            missing.append(name)
            continue
        pdf_bytes: bytes | None = None
        text = ""
        for chat_id in chat_ids:
            try:
                if send_pdf:
                    if pdf_bytes is None:
                        pdf_bytes = build_student_pdf_report(name, grouped[name], period)
                    caption = f"Отчет по марафону: {name}"
                    result = send_telegram_document(chat_id, pdf_bytes, report_filename(name), caption)
                else:
                    if not text:
                        text = build_student_telegram_report(name, grouped[name], period)
                    result = send_telegram_message(chat_id, text)
                sent.append({"name": name, "chatId": chat_id, "messageId": result.get("result", {}).get("message_id")})
                time.sleep(0.05)
            except Exception as error:
                errors.append({"name": name, "chatId": chat_id, "error": str(error)})

    return {
        "ok": len(errors) == 0,
        "format": "pdf" if send_pdf else "text",
        "sent": sent,
        "missing": missing,
        "errors": errors,
    }


def load_ratings(
    period_from: str,
    period_to: str,
    group_ids: set[int] | None,
    subjects: set[str] | None,
    include_virtual: bool,
    origins: set[str] | None,
    limit: int | None,
) -> dict[str, Any]:
    groups = selected_groups(
        group_ids=group_ids,
        subjects=subjects,
        include_virtual=include_virtual,
        origins=origins,
    )
    effective_limit = limit if limit is not None else MAX_GROUPS_PER_REQUEST
    if effective_limit > 0 and len(groups) > effective_limit:
        groups = groups[:effective_limit]

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    def load_group(group: GroupInfo):
        content = fetch_attendance_xlsx(group.id, period_from, period_to)
        return group, parse_attendance_xlsx(content, group)

    with ThreadPoolExecutor(max_workers=max(1, DEFAULT_CONCURRENCY)) as executor:
        futures = {executor.submit(load_group, group): group for group in groups}
        for future in as_completed(futures):
            group = futures[future]
            try:
                _, group_rows = future.result()
                rows.extend(group_rows)
            except Exception as error:
                errors.append({"groupId": group.id, "group": group.name, "error": str(error)})

    add_places(rows)

    return {
        "ok": True,
        "period": {"from": period_from, "to": period_to},
        "groups": [group.__dict__ for group in groups],
        "rows": rows,
        "errors": errors,
    }


def resolve_ratings_payload(query: dict[str, str]) -> dict[str, Any]:
    default_from, default_to = current_month_range()
    config = load_group_config()
    configured_ids, missing_names, missing_candidates = resolve_config_group_ids(fetch_group_tree(), config)
    period_from = query.get("periodFrom") or os.getenv("SOHOLMS_PERIOD_FROM") or config.get("periodFrom") or default_from
    period_to = query.get("periodTo") or os.getenv("SOHOLMS_PERIOD_TO") or config.get("periodTo") or default_to
    group_ids = parse_int_set(query.get("groupIds", "")) or configured_ids or None
    subjects = parse_str_set(query.get("subjects", ""))
    include_virtual = query.get("includeVirtual") == "1" or bool(config.get("includeVirtual"))
    origins = parse_origin_set(query.get("origins", ""))
    limit = parse_limit(query.get("limit", ""))
    cache_key = f"ratings:{period_from}:{period_to}:{group_ids}:{subjects}:{include_virtual}:{origins}:{limit}"
    payload = cached(
        cache_key,
        DEFAULT_CACHE_SECONDS,
        lambda: load_ratings(period_from, period_to, group_ids, subjects, include_virtual, origins, limit),
    )
    if missing_names:
        payload = {
            **payload,
            "missingConfigNames": missing_names,
            "missingConfigCandidates": missing_candidates,
        }
    return payload


def parse_int_set(value: str) -> set[int] | None:
    if not value:
        return None
    result = {int(part) for part in re.split(r"[,\s]+", value.strip()) if part}
    return result or None


def parse_str_set(value: str) -> set[str] | None:
    if not value:
        return None
    result = {infer_subject(part.strip()) for part in value.split(",") if part.strip()}
    result.discard("без предмета")
    return result or None


def parse_origin_set(value: str) -> set[str] | None:
    if not value:
        return None
    allowed = {
        "manual": "ManualGroup",
        "manualgroup": "ManualGroup",
        "auto": "AutoGroup",
        "autogroup": "AutoGroup",
    }
    result = set()
    for part in value.split(","):
        key = part.strip().lower()
        if not key:
            continue
        result.add(allowed.get(key, part.strip()))
    return result or None


def parse_limit(value: str) -> int | None:
    if not value:
        return None
    try:
        return max(0, int(value))
    except ValueError:
        raise BackendError("limit must be a number", HTTPStatus.BAD_REQUEST)


def current_month_range() -> tuple[str, str]:
    today = date.today()
    start = today.replace(day=1)
    if today.month == 12:
        end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    return start.isoformat(), end.isoformat()


def json_bytes(payload: Any) -> bytes:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


def admin_key_matches(value: str) -> bool:
    return bool(BACKEND_ADMIN_KEY) and value.strip() == BACKEND_ADMIN_KEY


class Handler(BaseHTTPRequestHandler):
    server_version = "SoholmsMarathonBackend/1.0"

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.NO_CONTENT)
        self.send_cors_headers()
        self.end_headers()

    def do_POST(self):
        try:
            parsed = urlparse(self.path)
            query = {key: values[-1] for key, values in parse_qs(parsed.query).items()}

            if parsed.path == "/api/telegram/send-report":
                self.require_admin(query)
                body = self.read_json_body()
                payload = resolve_ratings_payload(query)
                result = send_telegram_reports(
                    payload.get("rows", []),
                    payload.get("period", {}),
                    normalize_text(body.get("studentName") if isinstance(body, dict) else ""),
                    send_pdf=bool(isinstance(body, dict) and body.get("format") == "pdf"),
                )
                return self.send_json(result, HTTPStatus.OK if result.get("ok") else HTTPStatus.BAD_GATEWAY)

            self.send_json({"ok": False, "error": "Not found"}, HTTPStatus.NOT_FOUND)
        except BackendError as error:
            self.send_json({"ok": False, "error": str(error)}, error.status)
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            query = {key: values[-1] for key, values in parse_qs(parsed.query).items()}

            if parsed.path == "/health":
                return self.send_json({"ok": True})

            if parsed.path == "/api/debug/auth":
                self.require_admin(query)
                return self.send_json({
                    "ok": True,
                    "apiToken": token_fingerprint("SOHOLMS_API_TOKEN"),
                    "excelToken": token_fingerprint("SOHOLMS_EXCEL_TOKEN"),
                    "fallbackToken": token_fingerprint("SOHOLMS_TOKEN"),
                    "cacheItems": len(_CACHE),
                })

            if parsed.path == "/api/cache/clear":
                self.require_admin(query)
                clear_cache()
                return self.send_json({"ok": True, "cacheItems": len(_CACHE)})

            if parsed.path == "/api/debug/xlsx":
                self.require_admin(query)
                group_id_value = query.get("groupId", "")
                period_from = query.get("periodFrom") or os.getenv("SOHOLMS_PERIOD_FROM") or load_group_config().get("periodFrom") or current_month_range()[0]
                period_to = query.get("periodTo") or os.getenv("SOHOLMS_PERIOD_TO") or load_group_config().get("periodTo") or current_month_range()[1]
                try:
                    group_id = int(group_id_value)
                except ValueError:
                    raise BackendError("groupId is required", HTTPStatus.BAD_REQUEST)
                content = fetch_attendance_xlsx(group_id, period_from, period_to)
                return self.send_json({
                    "ok": True,
                    "groupId": group_id,
                    "period": {"from": period_from, "to": period_to},
                    **inspect_attendance_xlsx(content),
                })

            if parsed.path == "/api/groups":
                search = query.get("search", "").strip()
                if search:
                    return self.send_json({
                        "ok": True,
                        "groups": search_groups(search),
                    })

                config = load_group_config()
                use_config = query.get("configured") == "1"
                configured_ids, missing_names, missing_candidates = resolve_config_group_ids(fetch_group_tree(), config) if use_config else (None, [], {})
                include_virtual = query.get("includeVirtual") == "1" or (use_config and bool(config.get("includeVirtual")))
                groups = selected_groups(
                    group_ids=parse_int_set(query.get("groupIds", "")) or configured_ids,
                    subjects=parse_str_set(query.get("subjects", "")),
                    include_virtual=include_virtual,
                    origins=parse_origin_set(query.get("origins", "")),
                )
                return self.send_json({
                    "ok": True,
                    "groups": [group.__dict__ for group in groups],
                    "missingConfigNames": missing_names,
                    "missingConfigCandidates": missing_candidates,
                })

            if parsed.path == "/api/ratings":
                payload = resolve_ratings_payload(query)
                if query.get("public") == "1":
                    payload = strip_for_public(payload)
                return self.send_json(payload)

            self.send_json({"ok": False, "error": "Not found"}, HTTPStatus.NOT_FOUND)
        except BackendError as error:
            self.send_json({"ok": False, "error": str(error)}, error.status)
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def send_cors_headers(self):
        self.send_header("Access-Control-Allow-Origin", os.getenv("CORS_ORIGIN", "*"))
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "content-type,authorization,x-admin-key")
        self.send_header("Access-Control-Allow-Private-Network", "true")

    def require_admin(self, query: dict[str, str]) -> None:
        if not BACKEND_ADMIN_KEY:
            return
        value = self.headers.get("x-admin-key", "") or query.get("adminKey", "")
        if not admin_key_matches(value):
            raise BackendError("Forbidden", HTTPStatus.FORBIDDEN)

    def read_json_body(self) -> dict[str, Any]:
        length = int(self.headers.get("content-length") or 0)
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        try:
            data = json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            raise BackendError("Invalid JSON body", HTTPStatus.BAD_REQUEST)
        return data if isinstance(data, dict) else {}

    def send_json(self, payload: Any, status: int = HTTPStatus.OK):
        body = json_bytes(payload)
        self.send_response(status)
        self.send_header("content-type", "application/json; charset=utf-8")
        self.send_header("content-length", str(len(body)))
        self.send_cors_headers()
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt: str, *args):
        sys.stderr.write("%s - %s\n" % (self.log_date_time_string(), fmt % args))


def main():
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8787"))
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"Soholms backend listening on http://{host}:{port}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
